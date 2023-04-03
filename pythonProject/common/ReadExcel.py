
#封装一个操作excel文件的类
# 实例化对像时，两个参数：1、excel的文件名  2、要操作的表单名
# 类里面写两个方法：
#     第一个方法读：返回结果是个列表，列表有多个元素，每个元素都是字典格式，代表一条测试用例
#     第二个方法写：传入参数：行号、列号，值
import openpyxl

class ReadExcel():
    def __init__(self, filename, wb):                  #初始化文件名和表单名
        self.filename = filename
        self.wb = wb

    def read_data(self):
        file = openpyxl.load_workbook(self.filename)   #打开表格文件
        table = file[self.wb]                          #打开表单
        data = list(table.rows)                        #取得表格所有行的数据并转成列表
        key = [i.value for i in data[0]]               #将第一行的元素转成字典的key,可以写成：key = []
                                                       #                                for i in data[0]:
                                                       #                                key.append(i.value)

        cases = []                                     #创建空列表来装用例
        for j in data[1:]:                             #除第一行外，遍历所有行
            values = [v.value for v in j]              #遍历每一行的元素，然后一行一个列表
            case = dict(zip(key,values))               #将第一行列表中元素与后面每一行的元素打包成字典，也就是一条用例
            cases.append(case)                         #将每一条用例添加到列表里面
        return cases

    def write_data(self, row, column, value):
        file = openpyxl.load_workbook(self.filename)
        table = file[self.wb]

        table.cell(row=row, column=column, value=value)    #对行号和列号对应的表格填写进value
        file.save(self.filename)                           #编写完表格后，保存数据到表里
        # print(table.cell(row=row, column=column).value)


#测试例子如下：
# filename = r'E:\pythonlearn\Class\date.xlsx'
# wb = 'register'
# real = ReadExcel(filename,wb)
# result = real.write_data(2,5,'pass')
# print(real.read_data())

# 第二个写法
# import openpyxl
#
# class ReadExecl:
#     def __init__(self,filename,wb):   #初始化execl文件名和表单名
#         self.filename = filename
#         self.wb = wb
#
#
#     def open_file(self):
#         self.fh = openpyxl.load_workbook(self.filename)   #加载execl文件
#         self.wbb = self.fh[self.wb]                             # 打开对应的表单
#
#
#     def read_data(self):
#         self.open_file()
#
#         datas = list(self.wb.rows)     # 获取表单中的所有单元格对像
#
#         k = [i.value for i in datas[0]]
#
#         cases = []
#         for items in datas[1:]:
#             v = [j.value for j in items]
#             cases.append(dict(zip(k,v)))
#
#         return cases
#
#     def write_data(self,row,column,value):
#         self.open_file()
#
#         self.wbb.cell(row = row,column = column,value = value)  #往表单中指定行,指定列中写数据,这里变化是行号,列号是固定的
#         self.fh.save(self.filename)          #写完之后,要调用save方法保存到execl文件中


